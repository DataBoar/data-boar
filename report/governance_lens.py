"""
Governance Lens — map technical findings to GRC control-gap narratives (Pro tier).

Heuristic only; not legal advice. See docs/plans/PLAN_GOVERNANCE_LENS.md.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from config.governance_map_loader import (
    DEFAULT_GOVERNANCE_MAP_FILE,
    load_governance_map_entries,
    pattern_name_matches,
    resolve_governance_map_path,
    target_context_matches,
)

_logger = logging.getLogger(__name__)

ENTERPRISE_TIER_WARNING = "Enterprise framework requires enterprise license tier"
DEFAULT_ENTERPRISE_MAP_FILE = "config/governance_framework_map_enterprise.example.yaml"
LICENSED_ENTERPRISE_MAP_FILE = "config/governance_framework_map_enterprise.yaml"

RiskLevelPt = Literal["alto", "medio", "baixo"]

_NONPROD_TARGET_RE = re.compile(
    r"(nonprod|non-prod|homolog|staging|\bdev\b|\blab\b|test|uat|sandbox)",
    re.IGNORECASE,
)
_SENS_RANK = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}


@dataclass(frozen=True)
class ControlGap:
    framework_id: str
    framework_name: str
    control_gap_title: str
    control_gap_body: str
    recommendation: str
    deadline_days: int | None
    finding_count: int
    max_sensitivity: str
    sample_patterns: tuple[str, ...] = ()


@dataclass
class GovernanceLensResult:
    control_gaps: list[ControlGap] = field(default_factory=list)
    framework_summary: dict[str, int] = field(default_factory=dict)
    risk_level: RiskLevelPt = "baixo"


def split_pattern_detected(pattern_detected: str | None) -> list[str]:
    if not pattern_detected:
        return []
    return [p.strip() for p in str(pattern_detected).split(",") if p.strip()]


def infer_target_context(
    row: dict[str, Any],
    source: Literal["database", "filesystem", "application"],
) -> str:
    if source == "filesystem":
        return "filesystem"
    if source == "application":
        return "api"
    target_name = str(row.get("target_name") or "")
    if _NONPROD_TARGET_RE.search(target_name):
        return "database_nonprod"
    return "database"


def _max_sensitivity(a: str, b: str) -> str:
    ra = _SENS_RANK.get((a or "").upper(), 0)
    rb = _SENS_RANK.get((b or "").upper(), 0)
    if rb > ra:
        return (b or "LOW").upper()
    return (a or "LOW").upper()


def governance_lens_feature_allowed(config: dict[str, Any] | None) -> bool:
    """True when governance is enabled and the active license tier allows Pro lens."""
    cfg = config or {}
    gov = cfg.get("governance") or {}
    if not gov.get("enabled"):
        return False
    from core.licensing.guard import get_license_guard

    guard = get_license_guard(cfg)
    tier = str(gov.get("tier") or "pro").strip().lower()
    if tier == "enterprise":
        return guard.is_allowed("governance_lens_enterprise") or guard.is_allowed(
            "governance_lens_pro"
        )
    return guard.is_allowed("governance_lens_pro")


def enterprise_lens_enabled(config: dict[str, Any] | None) -> bool:
    """True when config asks for Enterprise modules and the license allows them."""
    cfg = config or {}
    gov = cfg.get("governance") or {}
    if str(gov.get("tier") or "").strip().lower() != "enterprise":
        return False
    from core.licensing.guard import get_license_guard

    return get_license_guard(cfg).is_allowed("governance_lens_enterprise")


def _looks_like_enterprise_map_file(map_file: str) -> bool:
    name = Path(map_file).name.lower()
    return (
        name
        in {
            "governance_framework_map_enterprise.yaml",
            "governance_framework_map_enterprise.example.yaml",
        }
        or "enterprise" in name
    )


def _framework_tier_allowed(fw_tier: str, config: dict[str, Any]) -> bool:
    from core.licensing.guard import get_license_guard

    guard = get_license_guard(config)
    ft = (fw_tier or "pro").strip().lower()
    if ft == "enterprise":
        return guard.is_allowed("governance_lens_enterprise")
    return guard.is_allowed("governance_lens_pro")


class GovernanceLensGenerator:
    """Build Governance Lens gaps from session findings + framework map YAML."""

    def __init__(
        self,
        config: dict[str, Any] | None,
        map_entries: list[dict[str, Any]] | None = None,
        *,
        config_path: Path | str | None = None,
    ) -> None:
        self.config = config or {}
        self._config_path = config_path
        self._enterprise_warned = False
        if map_entries is not None:
            self._entries = map_entries
        else:
            self._entries = self._load_map_entries()

    def _warn_enterprise_tier(self) -> None:
        if self._enterprise_warned:
            return
        _logger.warning(ENTERPRISE_TIER_WARNING)
        self._enterprise_warned = True

    def _load_map_entries(self) -> list[dict[str, Any]]:
        gov = self.config.get("governance") or {}
        map_file = str(gov.get("map_file") or DEFAULT_GOVERNANCE_MAP_FILE).strip()
        map_path = resolve_governance_map_path(map_file, self._config_path)
        entries = load_governance_map_entries(map_path)

        enterprise_map_file = str(gov.get("enterprise_map_file") or "").strip()
        wants_enterprise = bool(enterprise_map_file) or _looks_like_enterprise_map_file(
            map_file
        )
        if not enterprise_lens_enabled(self.config):
            if wants_enterprise:
                self._warn_enterprise_tier()
            return entries

        ent_file = enterprise_map_file or LICENSED_ENTERPRISE_MAP_FILE
        ent_path = resolve_governance_map_path(ent_file, self._config_path)
        if not ent_path.is_file() and not enterprise_map_file:
            ent_path = resolve_governance_map_path(
                DEFAULT_ENTERPRISE_MAP_FILE, self._config_path
            )
        if ent_path.is_file() and ent_path.resolve() != map_path.resolve():
            entries = entries + load_governance_map_entries(ent_path)
        return entries

    def generate_from_rows(
        self,
        db_rows: list[dict[str, Any]],
        fs_rows: list[dict[str, Any]],
        app_rows: list[dict[str, Any]] | None = None,
    ) -> GovernanceLensResult:
        aggregated: dict[tuple[str, str], dict[str, Any]] = {}
        nonprod_db_count = 0

        def _consume(
            row: dict[str, Any],
            source: Literal["database", "filesystem", "application"],
        ) -> None:
            nonlocal nonprod_db_count
            ctx = infer_target_context(row, source)
            if source == "database" and ctx == "database_nonprod":
                nonprod_db_count += 1
            sensitivity = str(row.get("sensitivity_level") or "LOW").upper()
            for pattern in split_pattern_detected(row.get("pattern_detected")):
                for entry in self._entries:
                    if not pattern_name_matches(entry["pattern_name"], pattern):
                        continue
                    if not target_context_matches(entry.get("target_context"), ctx):
                        continue
                    for fw in entry["frameworks"]:
                        if not _framework_tier_allowed(
                            fw.get("tier", "pro"), self.config
                        ):
                            if (
                                str(fw.get("tier") or "").strip().lower()
                                == "enterprise"
                            ):
                                self._warn_enterprise_tier()
                            continue
                        key = (fw["id"], fw["control_gap_title"])
                        bucket = aggregated.get(key)
                        if bucket is None:
                            bucket = {
                                "framework_id": fw["id"],
                                "framework_name": fw["name"],
                                "control_gap_title": fw["control_gap_title"],
                                "control_gap_body": fw["control_gap_body"],
                                "recommendation": fw["recommendation"],
                                "deadline_days": fw.get("deadline_days"),
                                "finding_count": 0,
                                "max_sensitivity": "LOW",
                                "patterns": set(),
                            }
                            aggregated[key] = bucket
                        bucket["finding_count"] += 1
                        bucket["max_sensitivity"] = _max_sensitivity(
                            bucket["max_sensitivity"], sensitivity
                        )
                        bucket["patterns"].add(pattern)

        for row in db_rows:
            _consume(row, "database")
        for row in fs_rows:
            _consume(row, "filesystem")
        for row in app_rows or []:
            _consume(row, "application")

        gaps: list[ControlGap] = []
        framework_summary: dict[str, int] = {}
        for bucket in aggregated.values():
            gap = ControlGap(
                framework_id=bucket["framework_id"],
                framework_name=bucket["framework_name"],
                control_gap_title=bucket["control_gap_title"],
                control_gap_body=bucket["control_gap_body"],
                recommendation=bucket["recommendation"],
                deadline_days=bucket["deadline_days"],
                finding_count=int(bucket["finding_count"]),
                max_sensitivity=bucket["max_sensitivity"],
                sample_patterns=tuple(sorted(bucket["patterns"]))[:5],
            )
            gaps.append(gap)
            framework_summary[gap.framework_id] = (
                framework_summary.get(gap.framework_id, 0) + 1
            )

        def _sort_key(g: ControlGap) -> tuple[int, int, str]:
            return (
                -_SENS_RANK.get(g.max_sensitivity, 0),
                -g.finding_count,
                g.framework_id,
            )

        gaps.sort(key=_sort_key)
        risk = _derive_risk_level(gaps, nonprod_db_count)
        return GovernanceLensResult(
            control_gaps=gaps,
            framework_summary=framework_summary,
            risk_level=risk,
        )

    def generate_from_session(
        self,
        db_manager: Any,
        session_id: str,
    ) -> GovernanceLensResult:
        db_rows, fs_rows, app_rows, _fail = db_manager.get_findings(session_id)
        return self.generate_from_rows(db_rows, fs_rows, app_rows)


def _derive_risk_level(
    gaps: list[ControlGap],
    nonprod_db_count: int,
) -> RiskLevelPt:
    if not gaps:
        return "baixo"
    if nonprod_db_count >= 3:
        return "alto"
    if any(g.max_sensitivity == "HIGH" for g in gaps):
        return "alto"
    return "medio"


def governance_result_to_excel_rows(
    result: GovernanceLensResult,
) -> list[dict[str, str]]:
    """Flatten result for the Governance View worksheet (pandas-friendly)."""
    cols = {
        "Seção": "",
        "Framework": "",
        "Controle ID": "",
        "Título": "",
        "Descrição": "",
        "Recomendação": "",
        "Prazo (dias)": "",
        "Qtd findings": "",
        "Severidade máx.": "",
    }
    rows: list[dict[str, str]] = []
    rows.append(
        {
            **cols,
            "Seção": "Aviso",
            "Framework": (
                "Governance Lens — heurístico; não é parecer jurídico nem conclusão de auditoria."
            ),
        }
    )
    rows.append({**cols, "Seção": "Nível de risco", "Framework": result.risk_level})
    rows.append({**cols})
    for gap in result.control_gaps:
        deadline = "" if gap.deadline_days is None else str(gap.deadline_days)
        rows.append(
            {
                **cols,
                "Framework": gap.framework_name,
                "Controle ID": gap.framework_id,
                "Título": gap.control_gap_title,
                "Descrição": gap.control_gap_body,
                "Recomendação": gap.recommendation,
                "Prazo (dias)": deadline,
                "Qtd findings": str(gap.finding_count),
                "Severidade máx.": gap.max_sensitivity,
            }
        )
    if result.framework_summary:
        rows.append({**cols})
        rows.append({**cols, "Seção": "Resumo por framework"})
        for fw_id, count in sorted(result.framework_summary.items()):
            rows.append({**cols, "Framework": fw_id, "Controle ID": str(count)})
    return rows


def apply_governance_risk_level_style(
    writer: Any, sheet_name: str, risk_level: str
) -> None:
    """Color the risk-level cell after the sheet is written."""
    from openpyxl.styles import PatternFill

    fills = {
        "alto": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "medio": PatternFill(fill_type="solid", fgColor="FFEB9C"),
        "baixo": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    }
    ws = writer.sheets.get(sheet_name)
    if ws is None:
        return
    fill = fills.get(risk_level, fills["baixo"])
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Nível de risco":
            row[1].fill = fill
            break
