"""Regression coverage for XLSX formula-injection safety across all writers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine

REPO_ROOT = Path(__file__).resolve().parents[1]
_DANGEROUS_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _discover_xlsx_writer_sources() -> set[str]:
    """Return repo source files that call DataFrame.to_excel."""
    found: set[str] = set()
    for path in REPO_ROOT.rglob("*.py"):
        rel = path.relative_to(REPO_ROOT)
        if rel.parts and rel.parts[0] in {"tests", ".venv", ".git"}:
            continue
        if b"to_excel(" in path.read_bytes():
            found.add(rel.as_posix())
    return found


def _assert_workbook_formula_safe(path: Path, writer_id: str) -> None:
    """Assert workbook has no string cells starting with formula-leading chars."""
    wb = load_workbook(path)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str) and value:
                        assert value[0] not in _DANGEROUS_FORMULA_PREFIXES, (
                            f"{writer_id}: unsafe formula-leading cell in "
                            f"{path.name}/{ws.title}: {value!r}"
                        )
    finally:
        wb.close()


def _run_report_generator_writer(tmp_path: Path) -> Path:
    from core.database import LocalDBManager
    from report.generator import generate_report

    db_path = tmp_path / "audit.db"
    out_dir = tmp_path / "out"
    out_dir.mkdir(parents=True, exist_ok=True)
    mgr = LocalDBManager(str(db_path))
    try:
        sid = "xlsx-safe-sid"
        mgr.set_current_session_id(sid)
        mgr.create_session_record(
            sid,
            tenant_name="=tenant-formula",
            technician_name="+technician-formula",
        )
        mgr.save_finding(
            "database",
            target_name="=db-target",
            column_name="+db-column",
            sensitivity_level="HIGH",
            pattern_detected="-db-pattern",
            norm_tag="@db-norm",
            ml_confidence=90,
        )
        mgr.save_finding(
            "filesystem",
            target_name="-fs-target",
            path="=fs/path",
            file_name="@fs-file.txt",
            sensitivity_level="HIGH",
            pattern_detected="+fs-pattern",
            norm_tag="-fs-norm",
            ml_confidence=88,
        )
        mgr.save_failure("=failed-target", "error", "@failure-details")
        mgr.save_data_source_inventory(
            target_name="+inventory-target",
            source_type="database",
            product="-postgres",
            raw_details="=inventory-details",
        )
        mgr.finish_session(sid)
        path = generate_report(mgr, sid, output_dir=str(out_dir), config={})
        assert path is not None
        return Path(path)
    finally:
        mgr.dispose()


def _run_grc_multiformat_writer(tmp_path: Path) -> Path:
    from report.grc_export_multiformat import export_grc_xlsx

    data = {
        "schema_version": "data_boar_grc_executive_report_v1",
        "report_metadata": {},
        "executive_summary": {},
        "detailed_findings": [
            {
                "asset_id": "=asset-formula",
                "asset_class": "+asset-class",
                "data_category": "-category",
                "risk_score": 8.8,
                "remediation_priority": "@priority",
                "regulatory_impact": "=impact",
                "location_summary": "+location",
                "violation_desc": "-desc",
                "norm_tags": ["@norm"],
                "pii_types": [
                    {"type": "=CPF", "count": 1, "exposure": "+cleartext"},
                ],
            }
        ],
        "recommendations": [],
    }
    path = tmp_path / "grc.xlsx"
    export_grc_xlsx(data, path)
    return path


def _run_utils_report_gen_writer(tmp_path: Path) -> Path:
    from utils.report_gen import ReportGenerator

    db_path = tmp_path / "legacy-audit.sqlite"
    engine = create_engine(f"sqlite:///{db_path}")
    df = pd.DataFrame(
        [
            {
                "target_name": "=legacy-target",
                "sensitivity_level": "HIGH",
                "timestamp": "2026-07-12T00:00:00Z",
                "pattern_detected": "+legacy-pattern",
            },
            {
                "target_name": "@legacy-target-2",
                "sensitivity_level": "MEDIUM",
                "timestamp": "2026-07-12T00:00:01Z",
                "pattern_detected": "-legacy-pattern-2",
            },
        ]
    )
    df.to_sql("audit_findings", engine, index=False, if_exists="replace")

    class _DbManagerStub:
        def __init__(self, db_engine):
            self.engine = db_engine

    generator = ReportGenerator(_DbManagerStub(engine))
    generator._create_heatmap = lambda _df: None
    before = set(tmp_path.glob("Relatorio_Compliance_*.xlsx"))
    cwd = Path.cwd()
    try:
        import os

        os.chdir(tmp_path)
        generator.generate_comprehensive_report()
    finally:
        os.chdir(cwd)
        engine.dispose()
    after = set(tmp_path.glob("Relatorio_Compliance_*.xlsx"))
    created = sorted(after - before)
    assert created, "utils.report_gen writer did not produce XLSX output"
    return created[-1]


def _run_scanners_report_generator_writer(tmp_path: Path) -> Path:
    from scanners.report_generator import ReportGenerator

    output = tmp_path / "scanner-legacy.xlsx"
    data = [
        {
            "table_name": "=table-formula",
            "column_name": "+column-formula",
            "sensitivity_score": 99,
        }
    ]
    cwd = Path.cwd()
    try:
        import os

        os.chdir(tmp_path)
        ReportGenerator(data).generate_report(str(output))
    finally:
        os.chdir(cwd)
    return output


_XLSX_WRITER_CASES = [
    (
        "report.generator.generate_report",
        "report/generator.py",
        _run_report_generator_writer,
    ),
    (
        "report.grc_export_multiformat.export_grc_xlsx",
        "report/grc_export_multiformat.py",
        _run_grc_multiformat_writer,
    ),
    (
        "utils.report_gen.ReportGenerator.generate_comprehensive_report",
        "utils/report_gen.py",
        _run_utils_report_gen_writer,
    ),
    (
        "scanners.report_generator.ReportGenerator.generate_report",
        "scanners/report_generator.py",
        _run_scanners_report_generator_writer,
    ),
]


def test_xlsx_writer_inventory_matches_repo_sources() -> None:
    """Guardrail: every repo XLSX writer is covered by this parametrized regression."""
    expected_sources = {case[1] for case in _XLSX_WRITER_CASES}
    discovered_sources = _discover_xlsx_writer_sources()
    assert discovered_sources == expected_sources


@pytest.mark.parametrize(
    ("writer_id", "source_file", "runner"),
    _XLSX_WRITER_CASES,
    ids=[case[0] for case in _XLSX_WRITER_CASES],
)
def test_xlsx_writers_neutralize_formula_prefixes(
    tmp_path: Path,
    writer_id: str,
    source_file: str,
    runner,
) -> None:
    """All XLSX writers must neutralize formula-leading characters in output cells."""
    assert source_file in _discover_xlsx_writer_sources()
    output_path = runner(tmp_path)
    assert output_path.is_file() and output_path.stat().st_size > 0
    _assert_workbook_formula_safe(output_path, writer_id)
