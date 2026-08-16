"""Phase 2a/2c: crypto_controls_audit persistence + connector wiring + report sheet."""

from __future__ import annotations

import sqlite3
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

from connectors.dataverse_connector import DataverseConnector
from connectors.mongodb_connector import MongoDBConnector
from connectors.powerbi_connector import PowerBIConnector
from connectors.redis_connector import RedisConnector
from connectors.rest_connector import RESTConnector
from connectors.smb_connector import SMBConnector
from connectors.sql_connector import SQLConnector
from connectors.url_guard import OPT_IN_KEY
from core.crypto_audit import collect_sql_crypto_facts
from core.database import LocalDBManager
from report.generator import generate_report


def _mock_httpx_tls_client(base_url: str = "https://example.com") -> MagicMock:
    """httpx.Client mock with TLS socket extras for collect_httpx_crypto_facts."""
    sock = MagicMock()
    sock.version.return_value = "TLSv1.3"
    sock.cipher.return_value = ("TLS_AES_256_GCM_SHA384", "TLSv1.3", 256)
    net = MagicMock()
    net.get_extra_info.side_effect = lambda name: sock if name == "socket" else None
    stream_resp = MagicMock()
    stream_resp.extensions = {"network_stream": net}
    stream_resp.__enter__.return_value = stream_resp
    stream_resp.__exit__.return_value = False
    client = MagicMock()
    client.base_url = base_url
    client.stream.return_value = stream_resp
    get_resp = MagicMock()
    get_resp.raise_for_status = MagicMock()
    get_resp.json.return_value = {"value": [], "ok": True}
    get_resp.text = "{}"
    client.get.return_value = get_resp
    return client


def test_crypto_controls_audit_save_and_get(tmp_path: Path) -> None:
    db_path = str(tmp_path / "crypto_audit.db")
    mgr = LocalDBManager(db_path)
    try:
        mgr.set_current_session_id("crypto-sess-1")
        mgr.create_session_record("crypto-sess-1")
        mgr.save_crypto_controls_audit(
            target_name="pg-main",
            connection_type="database",
            strong_crypto_result="ok",
            strong_crypto_details="source=pg_stat_ssl; tls=TLSv1.3",
        )
        rows = mgr.get_crypto_controls_audit("crypto-sess-1")
        assert len(rows) == 1
        assert rows[0]["target_name"] == "pg-main"
        assert rows[0]["strong_crypto_result"] == "ok"
        assert "TLSv1.3" in (rows[0]["strong_crypto_details"] or "")
        assert rows[0]["inferred_controls_summary"] is None
    finally:
        mgr.dispose()


def test_crypto_controls_audit_table_created_on_init(tmp_path: Path) -> None:
    """LocalDBManager ensures crypto_controls_audit exists."""
    from sqlalchemy import text

    db_path = str(tmp_path / "legacy.db")
    mgr = LocalDBManager(db_path)
    try:
        with mgr.engine.connect() as c:
            names = {
                row[0]
                for row in c.execute(
                    text("SELECT name FROM sqlite_master WHERE type='table'")
                )
            }
        assert "crypto_controls_audit" in names
    finally:
        mgr.dispose()


def test_collect_sql_crypto_facts_sqlite() -> None:
    from sqlalchemy import create_engine

    eng = create_engine("sqlite:///:memory:")
    facts = collect_sql_crypto_facts(eng, {"name": "local"})
    assert facts.source == "sqlite"


def test_collect_sql_crypto_facts_postgres_pg_stat_ssl() -> None:
    """Mock SQLAlchemy connect/execute to simulate pg_stat_ssl row."""

    class _FakeResult:
        def fetchone(self):
            return (True, "TLSv1.3", "ECDHE-RSA-AES256-GCM-SHA384")

    class _FakeConn:
        def execute(self, *_a, **_k):
            return _FakeResult()

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

    class _FakeDialect:
        name = "postgresql"

    class _FakeEngine:
        dialect = _FakeDialect()

        def connect(self):
            return _FakeConn()

    facts = collect_sql_crypto_facts(
        _FakeEngine(), {"sslmode": "verify-full", "name": "pg"}
    )
    assert facts.source == "pg_stat_ssl"
    assert facts.tls_in_use is True
    assert facts.tls_version == "TLSv1.3"
    assert facts.sslmode == "verify-full"


def test_sql_connector_saves_crypto_audit_when_flag_on(tmp_path: Path) -> None:
    db_path = tmp_path / "scan.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("CREATE TABLE t1 (a TEXT)")
    conn.execute("INSERT INTO t1 VALUES ('x')")
    conn.commit()
    conn.close()

    target = {
        "type": "database",
        "driver": "sqlite",
        "database": str(db_path),
        "name": "SQLiteCrypto",
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    connector = SQLConnector(target, scanner, db_manager)
    connector.run()

    assert db_manager.save_crypto_controls_audit.called
    kwargs = db_manager.save_crypto_controls_audit.call_args.kwargs
    assert kwargs["target_name"] == "SQLiteCrypto"
    assert kwargs["connection_type"] == "database"
    assert kwargs["strong_crypto_result"] == "not_applicable"


def test_sql_crypto_probe_exception_does_not_abort_scan(tmp_path: Path) -> None:
    """Phase 4.1: crypto collect/evaluate/persist errors must not stop sampling."""
    db_path = tmp_path / "scan_failsoft.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("CREATE TABLE t1 (a TEXT)")
    conn.execute("INSERT INTO t1 VALUES ('x')")
    conn.commit()
    conn.close()

    target = {
        "type": "database",
        "driver": "sqlite",
        "database": str(db_path),
        "name": "SQLiteFailSoft",
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "HIGH",
        "pattern_detected": "EMAIL",
        "norm_tag": "",
        "ml_confidence": 90,
    }
    db_manager = MagicMock()
    connector = SQLConnector(target, scanner, db_manager)
    with patch(
        "connectors.sql_connector.collect_sql_crypto_facts",
        side_effect=RuntimeError("probe boom"),
    ):
        connector.run()

    assert not db_manager.save_crypto_controls_audit.called
    assert scanner.scan_column.called
    assert db_manager.save_finding.called
    assert not db_manager.save_failure.called


def test_sql_connector_skips_crypto_audit_when_flag_off(tmp_path: Path) -> None:
    db_path = tmp_path / "scan_off.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("CREATE TABLE t1 (a TEXT)")
    conn.commit()
    conn.close()

    target = {
        "type": "database",
        "driver": "sqlite",
        "database": str(db_path),
        "name": "SQLiteOff",
        "_validate_crypto": False,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    connector = SQLConnector(target, scanner, db_manager)
    connector.run()

    assert not db_manager.save_crypto_controls_audit.called


def test_sql_connector_updates_inferred_controls_when_flag_on(tmp_path: Path) -> None:
    secret_col = "customer_cpf_hash"
    db_path = tmp_path / "scan_infer.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute(
        f'CREATE TABLE t1 (email TEXT, "{secret_col}" TEXT, email_masked TEXT)'
    )
    conn.execute(
        "INSERT INTO t1 VALUES (?, ?, ?)",
        ("a@example.com", "should-not-appear", "m***"),
    )
    conn.commit()
    conn.close()

    target = {
        "type": "database",
        "driver": "sqlite",
        "database": str(db_path),
        "name": "SQLiteInfer",
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    connector = SQLConnector(target, scanner, db_manager)
    connector.run()

    assert db_manager.save_crypto_controls_audit.called
    assert db_manager.update_crypto_controls_inferred_summary.called
    args = db_manager.update_crypto_controls_inferred_summary.call_args
    assert (
        args.args[0] == "SQLiteInfer" or args.kwargs.get("target_name") == "SQLiteInfer"
    )
    summary = (
        args.args[1]
        if len(args.args) > 1
        else args.kwargs.get("inferred_controls_summary")
    )
    assert summary
    assert "hashing" in summary
    assert "masking" in summary
    assert secret_col not in summary
    assert "cpf" not in summary
    assert "should-not-appear" not in summary
    assert "a@example.com" not in summary


def test_sql_inferred_controls_saved_when_mid_loop_raises(tmp_path: Path) -> None:
    """Bugbot: mid-loop exceptions must not skip Phase 3 inferred summary."""
    db_path = tmp_path / "scan_infer_boom.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute(
        "CREATE TABLE t1 (email TEXT, customer_cpf_hash TEXT, email_masked TEXT)"
    )
    conn.execute("INSERT INTO t1 VALUES ('a@x.com', 'h', 'm')")
    conn.commit()
    conn.close()

    target = {
        "type": "database",
        "driver": "sqlite",
        "database": str(db_path),
        "name": "SQLiteInferBoom",
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    db_manager = MagicMock()
    connector = SQLConnector(target, scanner, db_manager)

    def _boom(*_a, **_k):
        _boom.n = getattr(_boom, "n", 0) + 1
        if _boom.n >= 2:
            raise RuntimeError("mid-loop boom")

    with patch.object(SQLConnector, "_process_one_finding", side_effect=_boom):
        connector.run()

    assert db_manager.save_crypto_controls_audit.called
    assert db_manager.save_failure.called
    assert db_manager.update_crypto_controls_inferred_summary.called
    summary = db_manager.update_crypto_controls_inferred_summary.call_args.args[1]
    assert "hashing" in summary
    assert "customer_cpf_hash" not in summary


def test_mongodb_inferred_controls_saved_when_mid_loop_raises() -> None:
    target = {
        "name": "mongo-infer-boom",
        "host": "localhost",
        OPT_IN_KEY: True,
        "port": 27017,
        "database": "test",
        "tls": True,
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.side_effect = RuntimeError("mid-loop boom")
    db_manager = MagicMock()
    coll = MagicMock()
    coll.find.return_value.limit.return_value = [
        {"email": "a@x.com", "customer_cpf_hash": "h", "email_masked": "m"}
    ]
    db = MagicMock()
    db.list_collection_names.return_value = ["people"]
    db.__getitem__.return_value = coll
    db.command.return_value = {"version": "7.0.0"}
    client = MagicMock()
    client.__getitem__.return_value = db
    connector = MongoDBConnector(target, scanner, db_manager)
    with patch("connectors.mongodb_connector.MongoClient", return_value=client):
        with patch("connectors.mongodb_connector._MONGO_AVAILABLE", True):
            connector.run()

    assert db_manager.save_crypto_controls_audit.called
    assert db_manager.save_failure.called
    assert db_manager.update_crypto_controls_inferred_summary.called
    summary = db_manager.update_crypto_controls_inferred_summary.call_args.args[1]
    assert "hashing" in summary
    assert "customer_cpf_hash" not in summary


def test_redis_inferred_controls_saved_when_mid_loop_raises() -> None:
    target = {
        "name": "redis-infer-boom",
        "host": "localhost",
        OPT_IN_KEY: True,
        "port": 6379,
        "tls": False,
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.side_effect = RuntimeError("mid-loop boom")
    db_manager = MagicMock()
    client = MagicMock()
    client.info.return_value = {"redis_version": "7.2.0"}
    client.scan_iter.return_value = iter(["user_hash", "session_token", "plain"])
    fake_redis = MagicMock()
    fake_redis.Redis.return_value = client
    connector = RedisConnector(target, scanner, db_manager)
    with patch("connectors.redis_connector._REDIS_AVAILABLE", True):
        with patch("connectors.redis_connector.redis", fake_redis):
            connector.run()

    assert db_manager.save_crypto_controls_audit.called
    assert db_manager.save_failure.called
    assert db_manager.update_crypto_controls_inferred_summary.called
    summary = db_manager.update_crypto_controls_inferred_summary.call_args.args[1]
    assert "hashing" in summary
    assert "tokenization" in summary
    assert "user_hash" not in summary


def test_update_crypto_controls_inferred_summary_persists(tmp_path: Path) -> None:
    db_path = str(tmp_path / "crypto_infer.db")
    mgr = LocalDBManager(db_path)
    try:
        mgr.set_current_session_id("infer-sess-1")
        mgr.create_session_record("infer-sess-1")
        mgr.save_crypto_controls_audit(
            target_name="pg-main",
            connection_type="database",
            strong_crypto_result="ok",
            strong_crypto_details="source=pg_stat_ssl; tls=TLSv1.3",
        )
        mgr.update_crypto_controls_inferred_summary(
            "pg-main",
            "2 names suggest hashing (heuristic; not verified — human review required)",
        )
        rows = mgr.get_crypto_controls_audit("infer-sess-1")
        assert len(rows) == 1
        assert "hashing" in (rows[0]["inferred_controls_summary"] or "")
    finally:
        mgr.dispose()


def test_mongodb_connect_honors_tls_flag() -> None:
    target = {
        "name": "mongo-tls",
        "host": "localhost",
        OPT_IN_KEY: True,
        "port": 27017,
        "database": "test",
        "tls": True,
    }
    connector = MongoDBConnector(target, MagicMock(), MagicMock())
    with patch("connectors.mongodb_connector.MongoClient") as mongo_mock:
        with patch("connectors.mongodb_connector._MONGO_AVAILABLE", True):
            mongo_mock.return_value = MagicMock()
            connector.connect()
    assert connector._tls_enabled is True
    assert mongo_mock.call_args.kwargs.get("tls") is True


def test_redis_connect_honors_tls_and_cert_reqs() -> None:
    target = {
        "name": "redis-tls",
        "host": "localhost",
        OPT_IN_KEY: True,
        "port": 6379,
        "tls": True,
        "ssl_cert_reqs": "required",
    }
    connector = RedisConnector(target, MagicMock(), MagicMock())
    fake_redis = MagicMock()
    fake_client = MagicMock()
    fake_redis.Redis.return_value = fake_client
    # Real ConnectionPool path: patch only redis.Redis factory; pool is built in-module.
    with patch("connectors.redis_connector._REDIS_AVAILABLE", True):
        with patch("connectors.redis_connector.redis", fake_redis):
            with patch(
                "connectors.url_guard._resolve_host_ips",
                return_value=[__import__("ipaddress").ip_address("127.0.0.1")],
            ):
                connector.connect()
    assert connector._tls_enabled is True
    assert fake_redis.Redis.called
    pool = fake_redis.Redis.call_args.kwargs.get("connection_pool")
    assert pool is not None
    from redis.connection import SSLConnection

    assert issubclass(pool.connection_class, SSLConnection)
    # Hostname kept for TLS SNI; cert_reqs forwarded into pool connection kwargs.
    assert pool.connection_kwargs.get("host") == "localhost"
    assert pool.connection_kwargs.get("ssl_cert_reqs") == "required"


def test_mongodb_connector_saves_crypto_audit_when_flag_on() -> None:
    target = {
        "name": "mongo-crypto",
        "host": "localhost",
        OPT_IN_KEY: True,
        "port": 27017,
        "database": "test",
        "tls": True,
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    client = MagicMock()
    client.__getitem__.return_value = MagicMock()
    client.__getitem__.return_value.list_collection_names.return_value = []
    client.__getitem__.return_value.command.return_value = {"version": "7.0.0"}
    connector = MongoDBConnector(target, scanner, db_manager)
    with patch("connectors.mongodb_connector.MongoClient", return_value=client):
        with patch("connectors.mongodb_connector._MONGO_AVAILABLE", True):
            connector.run()
    assert db_manager.save_crypto_controls_audit.called
    kwargs = db_manager.save_crypto_controls_audit.call_args.kwargs
    assert kwargs["target_name"] == "mongo-crypto"
    assert kwargs["connection_type"] == "mongodb"
    assert kwargs["strong_crypto_result"] in {
        "ok",
        "warning",
        "fail",
        "not_available",
        "not_applicable",
    }
    details = kwargs["strong_crypto_details"] or ""
    assert "password=" not in details.lower()


def test_redis_connector_saves_crypto_audit_when_flag_on() -> None:
    target = {
        "name": "redis-crypto",
        "host": "localhost",
        OPT_IN_KEY: True,
        "port": 6379,
        "tls": False,
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    client = MagicMock()
    client.info.return_value = {"redis_version": "7.2.0"}
    client.scan_iter.return_value = iter([])
    fake_redis = MagicMock()
    fake_redis.Redis.return_value = client
    connector = RedisConnector(target, scanner, db_manager)
    with patch("connectors.redis_connector._REDIS_AVAILABLE", True):
        with patch("connectors.redis_connector.redis", fake_redis):
            connector.run()
    assert db_manager.save_crypto_controls_audit.called
    kwargs = db_manager.save_crypto_controls_audit.call_args.kwargs
    assert kwargs["target_name"] == "redis-crypto"
    assert kwargs["connection_type"] == "redis"
    assert kwargs["strong_crypto_result"] == "fail"  # plaintext
    details = kwargs["strong_crypto_details"] or ""
    assert "password=" not in details.lower()


def test_smb_connect_honors_encrypt_and_signing() -> None:
    target = {
        "name": "smb-tls",
        "host": "files.example.com",
        "share": "data",
        "user": "u",
        "password": "p",
        "encrypt": True,
        "require_signing": True,
        "_validate_crypto": False,
    }
    scanner = MagicMock()
    db_manager = MagicMock()
    session = MagicMock()
    session.signing_required = True
    session.encrypt_data = True
    session.connection.dialect = 0x0311
    session.connection.supports_encryption = True
    fake_smb = MagicMock()
    fake_smb.register_session.return_value = session
    fake_smb.walk.return_value = []
    connector = SMBConnector(target, scanner, db_manager)
    with patch("connectors.smb_connector._SMB_AVAILABLE", True):
        with patch("connectors.smb_connector.smbclient", fake_smb):
            connector.run()
    kwargs = fake_smb.register_session.call_args.kwargs
    assert kwargs.get("encrypt") is True
    assert kwargs.get("require_signing") is True


def test_smb_connector_saves_crypto_audit_when_flag_on() -> None:
    secret = "must-not-appear-in-crypto-details"
    target = {
        "name": "smb-crypto",
        "host": "files.example.com",
        "share": "data",
        "user": "u",
        "password": secret,
        "encrypt": True,
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    db_manager = MagicMock()
    session = MagicMock()
    session.signing_required = True
    session.encrypt_data = True
    session.connection.dialect = 0x0311
    session.connection.supports_encryption = True
    fake_smb = MagicMock()
    fake_smb.register_session.return_value = session
    fake_smb.walk.return_value = []
    connector = SMBConnector(target, scanner, db_manager)
    with patch("connectors.smb_connector._SMB_AVAILABLE", True):
        with patch("connectors.smb_connector.smbclient", fake_smb):
            connector.run()
    assert db_manager.save_crypto_controls_audit.called
    kwargs = db_manager.save_crypto_controls_audit.call_args.kwargs
    assert kwargs["target_name"] == "smb-crypto"
    assert kwargs["connection_type"] == "smb"
    assert kwargs["strong_crypto_result"] == "ok"
    details = kwargs["strong_crypto_details"] or ""
    assert secret not in details
    assert "password=" not in details.lower()


def test_smb_connector_skips_crypto_audit_when_flag_off() -> None:
    target = {
        "name": "smb-off",
        "host": "files.example.com",
        "share": "data",
        "user": "u",
        "password": "p",
        "_validate_crypto": False,
    }
    scanner = MagicMock()
    db_manager = MagicMock()
    session = MagicMock()
    fake_smb = MagicMock()
    fake_smb.register_session.return_value = session
    fake_smb.walk.return_value = []
    connector = SMBConnector(target, scanner, db_manager)
    with patch("connectors.smb_connector._SMB_AVAILABLE", True):
        with patch("connectors.smb_connector.smbclient", fake_smb):
            connector.run()
    assert not db_manager.save_crypto_controls_audit.called


def test_rest_connector_honors_verify_and_saves_crypto_audit() -> None:
    secret = "must-not-appear-in-rest-crypto-details"
    target = {
        "name": "rest-crypto",
        "base_url": "https://example.com",
        "paths": ["/health"],
        "verify": False,
        "auth": {"type": "bearer", "token": secret},
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    client = _mock_httpx_tls_client("https://example.com")
    fake_httpx = MagicMock()
    fake_httpx.Timeout = MagicMock(return_value=MagicMock())
    fake_httpx.BasicAuth = MagicMock()
    connector = RESTConnector(target, scanner, db_manager)
    with patch("connectors.rest_connector._HTTPX_AVAILABLE", True):
        with patch("connectors.rest_connector.httpx", fake_httpx):
            with patch(
                "connectors.rest_connector.build_pinned_httpx_client",
                return_value=client,
            ) as mock_build:
                connector.run()
    assert mock_build.call_args.kwargs.get("verify") is False
    assert db_manager.save_crypto_controls_audit.called
    kwargs = db_manager.save_crypto_controls_audit.call_args.kwargs
    assert kwargs["target_name"] == "rest-crypto"
    assert kwargs["connection_type"] == "rest"
    details = kwargs["strong_crypto_details"] or ""
    assert secret not in details
    assert "Bearer" not in details
    assert "token=" not in details.lower()


def test_rest_crypto_probe_exception_does_not_abort_scan() -> None:
    """Phase 4.1: httpx crypto probe errors must not stop path sampling."""
    target = {
        "name": "rest-failsoft",
        "base_url": "https://example.com",
        "paths": ["/health"],
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "HIGH",
        "pattern_detected": "EMAIL",
        "norm_tag": "",
        "ml_confidence": 90,
    }
    db_manager = MagicMock()
    client = _mock_httpx_tls_client("https://example.com")
    client.get.return_value.json.return_value = {"email": "a@example.com"}
    fake_httpx = MagicMock()
    fake_httpx.Timeout = MagicMock(return_value=MagicMock())
    connector = RESTConnector(target, scanner, db_manager)
    with patch("connectors.rest_connector._HTTPX_AVAILABLE", True):
        with patch("connectors.rest_connector.httpx", fake_httpx):
            with patch(
                "connectors.rest_connector.build_pinned_httpx_client",
                return_value=client,
            ):
                with patch(
                    "connectors.rest_connector.collect_httpx_crypto_facts",
                    side_effect=RuntimeError("probe boom"),
                ):
                    connector.run()

    assert not db_manager.save_crypto_controls_audit.called
    assert client.get.called
    assert db_manager.save_finding.called
    failure_msgs = [
        str(c.args[2]) if len(c.args) > 2 else str(c.kwargs)
        for c in db_manager.save_failure.call_args_list
    ]
    assert not any("probe boom" in m for m in failure_msgs)


def test_rest_connector_skips_crypto_audit_when_flag_off() -> None:
    target = {
        "name": "rest-off",
        "base_url": "https://example.com",
        "paths": ["/health"],
        "_validate_crypto": False,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    client = _mock_httpx_tls_client()
    fake_httpx = MagicMock()
    fake_httpx.Timeout = MagicMock(return_value=MagicMock())
    connector = RESTConnector(target, scanner, db_manager)
    with patch("connectors.rest_connector._HTTPX_AVAILABLE", True):
        with patch("connectors.rest_connector.httpx", fake_httpx):
            with patch(
                "connectors.rest_connector.build_pinned_httpx_client",
                return_value=client,
            ):
                connector.run()
    assert not db_manager.save_crypto_controls_audit.called


def test_powerbi_connector_saves_crypto_audit_when_flag_on() -> None:
    secret = "pbi-client-secret-must-not-leak"
    target = {
        "name": "pbi-crypto",
        "tenant_id": "tenant",
        "client_id": "cid",
        "client_secret": secret,
        "verify_ssl": True,
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    db_manager = MagicMock()
    client = _mock_httpx_tls_client("https://api.powerbi.com/v1.0")
    fake_httpx = MagicMock()
    fake_httpx.Timeout = MagicMock(return_value=MagicMock())
    connector = PowerBIConnector(target, scanner, db_manager)
    with patch("connectors.powerbi_connector._HTTPX_AVAILABLE", True):
        with patch(
            "connectors.powerbi_connector._get_access_token",
            return_value="access-token",
        ):
            with patch("connectors.powerbi_connector.httpx", fake_httpx):
                with patch(
                    "connectors.powerbi_connector.build_pinned_httpx_client",
                    return_value=client,
                ) as mock_build:
                    connector.run()
    assert mock_build.call_args.kwargs.get("verify") is True
    assert db_manager.save_crypto_controls_audit.called
    kwargs = db_manager.save_crypto_controls_audit.call_args.kwargs
    assert kwargs["connection_type"] == "powerbi"
    details = kwargs["strong_crypto_details"] or ""
    assert secret not in details
    assert "access-token" not in details


def test_dataverse_connector_saves_crypto_audit_when_flag_on() -> None:
    secret = "dv-client-secret-must-not-leak"
    target = {
        "name": "dv-crypto",
        "org_url": "https://org.crm.dynamics.com",
        "tenant_id": "tenant",
        "client_id": "cid",
        "client_secret": secret,
        "verify": True,
        "_validate_crypto": True,
    }
    scanner = MagicMock()
    db_manager = MagicMock()
    client = _mock_httpx_tls_client("https://org.api.crm.dynamics.com/api/data/v9.2")
    fake_httpx = MagicMock()
    fake_httpx.Timeout = MagicMock(return_value=MagicMock())
    connector = DataverseConnector(target, scanner, db_manager)
    with patch("connectors.dataverse_connector._HTTPX_AVAILABLE", True):
        with patch(
            "connectors.dataverse_connector._dataverse_token", return_value="dv-token"
        ):
            with patch("connectors.dataverse_connector.httpx", fake_httpx):
                with patch(
                    "connectors.dataverse_connector.build_pinned_httpx_client",
                    return_value=client,
                ) as mock_build:
                    connector.run()
    assert mock_build.call_args.kwargs.get("verify") is True
    assert db_manager.save_crypto_controls_audit.called
    kwargs = db_manager.save_crypto_controls_audit.call_args.kwargs
    assert kwargs["connection_type"] == "dataverse"
    details = kwargs["strong_crypto_details"] or ""
    assert secret not in details
    assert "dv-token" not in details
    assert "crm.dynamics.com" not in details


def test_mongodb_connector_skips_crypto_audit_when_flag_off() -> None:
    target = {
        "name": "mongo-off",
        "host": "localhost",
        OPT_IN_KEY: True,
        "database": "test",
        "_validate_crypto": False,
    }
    scanner = MagicMock()
    scanner.scan_column.return_value = {
        "sensitivity_level": "LOW",
        "pattern_detected": "NONE",
        "norm_tag": "",
        "ml_confidence": 0,
    }
    db_manager = MagicMock()
    client = MagicMock()
    client.__getitem__.return_value = MagicMock()
    client.__getitem__.return_value.list_collection_names.return_value = []
    client.__getitem__.return_value.command.return_value = {"version": "7.0.0"}
    connector = MongoDBConnector(target, scanner, db_manager)
    with patch("connectors.mongodb_connector.MongoClient", return_value=client):
        with patch("connectors.mongodb_connector._MONGO_AVAILABLE", True):
            connector.run()
    assert not db_manager.save_crypto_controls_audit.called


def test_generate_report_includes_crypto_controls_sheet(tmp_path: Path) -> None:
    out_dir = str(tmp_path / "out")
    Path(out_dir).mkdir()
    db_path = str(tmp_path / "report.db")
    mgr = LocalDBManager(db_path)
    try:
        mgr.set_current_session_id("rep-crypto")
        mgr.create_session_record("rep-crypto")
        mgr.save_finding(
            "database",
            target_name="T1",
            column_name="email",
            sensitivity_level="LOW",
            pattern_detected="EMAIL",
            norm_tag="",
            ml_confidence=10,
        )
        mgr.save_crypto_controls_audit(
            target_name="pg1",
            connection_type="database",
            strong_crypto_result="warning",
            strong_crypto_details="source=config_sslmode; sslmode=require",
            inferred_controls_summary=(
                "2 names suggest hashing (heuristic; not verified — human review required)"
            ),
        )
        mgr.finish_session("rep-crypto")
        path = generate_report(mgr, "rep-crypto", output_dir=out_dir, config={})
        assert path
        wb = openpyxl.load_workbook(path)
        assert "Crypto & controls" in wb.sheetnames
        ws = wb["Crypto & controls"]
        values = [[c.value for c in row] for row in ws.iter_rows(min_row=1, max_row=4)]
        assert values[0][0] == "Target"
        assert any(row[0] == "(note)" for row in values[1:])
        note = next(row for row in values[1:] if row[0] == "(note)")
        assert "human review" in (note[4] or "").lower()
        assert "compliance" in (note[4] or "").lower()
        assert "data source inventory" in (note[3] or "").lower()
        assert any(row[0] == "pg1" for row in values[1:])
        assert any(row[2] == "warning" for row in values[1:])
        assert any("hashing" in (row[4] or "") for row in values[1:])
    finally:
        mgr.dispose()


def test_generate_report_omits_crypto_sheet_when_no_rows(tmp_path: Path) -> None:
    out_dir = str(tmp_path / "out2")
    Path(out_dir).mkdir()
    db_path = str(tmp_path / "report2.db")
    mgr = LocalDBManager(db_path)
    try:
        mgr.set_current_session_id("rep-empty")
        mgr.create_session_record("rep-empty")
        mgr.save_finding(
            "database",
            target_name="T1",
            column_name="email",
            sensitivity_level="LOW",
            pattern_detected="EMAIL",
            norm_tag="",
            ml_confidence=10,
        )
        mgr.finish_session("rep-empty")
        path = generate_report(mgr, "rep-empty", output_dir=out_dir, config={})
        assert path
        wb = openpyxl.load_workbook(path)
        assert "Crypto & controls" not in wb.sheetnames
    finally:
        mgr.dispose()
